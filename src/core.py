from enum import Enum

from openpyxl import load_workbook
from sympy import Symbol
from sympy.solvers import solve


# тип стены
class WallType(Enum):
    STRETCH = 1  # гибкая связь
    STRONG = 2  # жесткая связь


class Input:

    def __init__(self, city: str,
                 temperature: float,
                 humidity: float,
                 wall_type: WallType,
                 insulation_material: str,
                 width_print_layer: float,
                 width_construction_layer: float):
        self.city = city  # город
        self.temperature = temperature  # температура внутри помещения
        self.humidity = humidity  # относительная влажность внутри помещения
        self.wall_type = wall_type  # тип стены
        self.insulation_material = insulation_material  # утеплитель
        self.width_print_layer = width_print_layer  # толщина печатаемого слоя
        self.width_construction_layer = width_construction_layer  # толщина конструктивного слоя


def read_file(file_path: str) -> list:
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    rows = sheet.rows
    wb.close()
    return rows


def get_regions_list():
    regions = {}
    for row in list(read_file("../resources/cities.xlsx"))[1:]:
        region = row[0].value
        city = row[1].value
        count_days = row[2].value
        temperature = row[3].value
        if region not in regions:
            regions[region] = {"cities": []}
        regions[region]["cities"].append({"city": city,
                                          "count_days": count_days,
                                          "temperature": temperature})
    return regions


def get_materials_list():
    materials = {}
    for row in list(read_file("../resources/material.xlsx"))[1:]:
        material = row[0].value
        coefficient = row[1].value
        materials[material] = coefficient
    print(materials)
    return materials


def mm2m(mm):
    return mm / 1000


def calc_layer_1(i: Input):
    return mm2m(i.width_print_layer) / 1.5


def calc_layer_3(i: Input):
    return mm2m(i.width_print_layer) / 1.5


def calc_layer_4(i: Input):
    return mm2m(i.width_print_layer) / 1.5


def calc_layer_5(i: Input):
    return mm2m(i.width_construction_layer) / 1.5


def calc_layer_6(i: Input):
    return mm2m(i.width_print_layer) / 1.5


def calc_width_insulation_material(i: Input):
    temperature_out = -7.5
    count_days = 214
    lambda2 = 0.052

    gcop = (i.temperature - temperature_out) * count_days
    a = 0.00035
    b = 1.4
    r = a * gcop + b

    x = Symbol('x')
    res = solve(1 / 8.7 +
                calc_layer_1(i) +
                mm2m(x) / lambda2 +
                calc_layer_3(i) +
                calc_layer_4(i) +
                calc_layer_5(i) +
                calc_layer_6(i) +
                1 / 23 - r,
                x)
    return res[0]
